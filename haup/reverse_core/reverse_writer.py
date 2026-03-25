"""
File Summary:
Async writer thread that decouples parsing from I/O.
Buffers 2000 rows then flushes to PostgreSQL/Neon (INSERT … ON CONFLICT DO NOTHING)
or xlsx (ws.append).  Marks chunk DONE in checkpoint only AFTER confirmed write.

====================================================================
Startup
====================================================================

ReverseWriter()  [Class → Object]
||
├── __init__()  [Function] -------------------------------> Store queues, schema, target; select flush function
│       │
│       ├── [Conditional Branch] SQLTarget ----------------> Set _flush_fn to _flush_sql
│       │
│       └── [Conditional Branch] ExcelTarget -------------> _init_excel_workbook(), set _flush_fn to _flush_excel
│
├── start()  [Function] ----------------------------------> Launch background writer thread
│       │
│       └── _run()  [Function] ---------------------------> Main writer loop
│               │
│               ├── PostgreSQL setup ---------------------> Connect via psycopg2, call _init_sql_table()
│               │
│               ├── Drain result_q -----------------------> Get parsed row packets
│               │
│               ├── [Conditional Branch] packet.has_error -> mark_failed(), accumulate parse_fails
│               │
│               ├── Buffer rows (BUFFER_SIZE=2000) -------> Accumulate before flush
│               │
│               ├── [Conditional Branch] buffer full -----> Call _flush_fn(buffer, chunk_id)
│               │
│               └── [Early Exit Branch] stop + empty -----> Final flush of remaining buffer
│
├── stop()  [Function] -----------------------------------> Set stop event and join thread
│
├── _init_sql_table()  [Function] -----------------------> CREATE TABLE IF NOT EXISTS with schema
│
├── _flush_sql()  [Function] ----------------------------> Bulk INSERT to PostgreSQL/Neon
│       │
│       ├── INSERT … ON CONFLICT DO NOTHING -------------> Idempotent upsert
│       ├── cursor.executemany() ------------------------> Batch insert all buffered rows
│       ├── checkpoint.mark_done() ----------------------> Crash-safe chunk tracking
│       │
│       └── [Exception Block] ---------------------------> Log error, mark_failed, re-raise
│
├── _init_excel_workbook()  [Function] ------------------> Create openpyxl workbook and write header
│       │
│       ├── [Conditional Branch] total_rows > 500k -----> Use write_only mode for large files
│       │
│       └── wb.save() -----------------------------------> Verify path is writable
│
├── _flush_excel()  [Function] --------------------------> Append rows to worksheet and save
│       │
│       ├── ws.append() ---------------------------------> Add each row to sheet
│       ├── wb.save() -----------------------------------> Persist to disk
│       ├── checkpoint.mark_done() ----------------------> Track progress
│       │
│       └── [Exception Block] ---------------------------> Log error, mark_failed, re-raise
│
├── total_rows_written  [Property] ----------------------> Thread-safe row count
│
└── total_parse_fails  [Property] -----------------------> Thread-safe fail count

====================================================================
FUNCTION / CLASS ENTRY POINT MARKERS
====================================================================
"""

from __future__ import annotations

import logging
import queue
import threading
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)

BUFFER_SIZE               = 2_000
EXCEL_WRITEONLY_THRESHOLD = 500_000


"""================= Startup class SQLTarget ================="""
@dataclass
class SQLTarget:
    conn:       Any
    table_name: str
    dialect:    str = "postgresql"
"""================= End class SQLTarget ================="""


"""================= Startup class ExcelTarget ================="""
@dataclass
class ExcelTarget:
    output_path: str
    total_rows:  int = 0
"""================= End class ExcelTarget ================="""


"""================= Startup class ReverseWriter ================="""
class ReverseWriter:

    """================= Startup method __init__ ================="""
    def __init__(
        self,
        result_q:   queue.Queue,
        checkpoint,
        schema,
        target,
    ):
        self.result_q   = result_q
        self.checkpoint = checkpoint
        self.schema     = schema
        self.target     = target

        self._stop_event         = threading.Event()
        self._thread: Optional[threading.Thread] = None
        self._total_rows_written = 0
        self._total_parse_fails  = 0
        self._lock               = threading.Lock()

        if isinstance(target, SQLTarget):
            self._flush_fn = self._flush_sql
        elif isinstance(target, ExcelTarget):
            self._wb, self._ws = self._init_excel_workbook(target)
            self._flush_fn = self._flush_excel
        else:
            raise TypeError(f"Unknown target type: {type(target)}")
    """================= End method __init__ ================="""

    """================= Startup method start ================="""
    def start(self) -> None:
        self._thread = threading.Thread(
            target = self._run,
            name   = "haup-writer",
            daemon = True,
        )
        self._thread.start()
        logger.info("[ReverseWriter] writer thread started")
    """================= End method start ================="""

    """================= Startup method stop ================="""
    def stop(self, timeout: int = 60) -> None:
        self._stop_event.set()
        if self._thread:
            self._thread.join(timeout=timeout)
    """================= End method stop ================="""

    """================= Startup method total_rows_written ================="""
    @property
    def total_rows_written(self) -> int:
        with self._lock:
            return self._total_rows_written
    """================= End method total_rows_written ================="""

    """================= Startup method total_parse_fails ================="""
    @property
    def total_parse_fails(self) -> int:
        with self._lock:
            return self._total_parse_fails
    """================= End method total_parse_fails ================="""

    """================= Startup method _run ================="""
    def _run(self) -> None:
        if isinstance(self.target, SQLTarget):
            try:
                import psycopg2
                self.target.conn = psycopg2.connect(**self.target._pg_config)
                self._init_sql_table(self.target)
                logger.info("[ReverseWriter] PostgreSQL/Neon connection established and table created")
            except Exception as exc:
                logger.error("[ReverseWriter] Failed to initialize SQL target: %s", exc)
                import traceback
                traceback.print_exc()
                return

        buffer:        list[dict] = []
        last_chunk_id: int        = 0

        while not self._stop_event.is_set() or not self.result_q.empty():
            try:
                packet = self.result_q.get(timeout=0.5)
            except queue.Empty:
                if buffer:
                    self._flush_fn(buffer, last_chunk_id)
                    buffer = []
                continue

            if packet.has_error:
                self.checkpoint.mark_failed(packet.chunk_id)
                with self._lock:
                    self._total_parse_fails += packet.parse_fails
                continue

            with self._lock:
                self._total_parse_fails += packet.parse_fails

            last_chunk_id = packet.chunk_id

            for row in packet.rows:
                row["__orig_rowid__"] = row.pop("__rowid__", None)
                buffer.append(row)

            if len(buffer) >= BUFFER_SIZE:
                self._flush_fn(buffer, last_chunk_id)
                buffer = []

        if buffer:
            self._flush_fn(buffer, last_chunk_id)

        logger.info(
            "[ReverseWriter] done  rows_written=%d  parse_fails=%d",
            self._total_rows_written, self._total_parse_fails,
        )
    """================= End method _run ================="""

    """================= Startup method _init_sql_table ================="""
    def _init_sql_table(self, target: SQLTarget) -> None:
        if target.conn is None:
            return

        col_defs = self.schema.sql_col_defs()
        sql = f"""
            CREATE TABLE IF NOT EXISTS "{target.table_name}" (
                __extract_id   SERIAL PRIMARY KEY,
                {col_defs},
                __orig_rowid__ VARCHAR(255) UNIQUE
            )
        """
        cursor = target.conn.cursor()
        cursor.execute(sql)
        cursor.execute(
            f'CREATE INDEX IF NOT EXISTS idx_{target.table_name}_orig_rowid '
            f'ON "{target.table_name}" (__orig_rowid__)'
        )
        target.conn.commit()
        cursor.close()

        logger.info("[ReverseWriter] PostgreSQL table ready: %s", target.table_name)
    """================= End method _init_sql_table ================="""

    """================= Startup method _flush_sql ================="""
    def _flush_sql(self, buffer: list[dict], chunk_id: int) -> None:
        target     = self.target
        final_cols = self.schema.final_cols + ["__orig_rowid__"]

        col_str      = ", ".join(f'"{c}"' for c in final_cols)
        placeholders = ", ".join(["%s"] * len(final_cols))
        sql = (
            f'INSERT INTO "{target.table_name}" ({col_str}) '
            f'VALUES ({placeholders}) '
            f'ON CONFLICT (__orig_rowid__) DO NOTHING'
        )

        values = [
            [row.get(c) for c in final_cols]
            for row in buffer
        ]

        try:
            cursor = target.conn.cursor()
            cursor.executemany(sql, values)
            rows_affected = cursor.rowcount
            target.conn.commit()
            cursor.close()

            n = rows_affected if rows_affected >= 0 else len(buffer)
            with self._lock:
                self._total_rows_written += n
            self.checkpoint.mark_done(chunk_id, rows_done=n)
            logger.debug(
                "[ReverseWriter] SQL flush  chunk=%d  rows=%d  affected=%d",
                chunk_id, len(buffer), n,
            )
        except Exception as exc:
            logger.error("[ReverseWriter] SQL flush error chunk=%d: %s", chunk_id, exc)
            self.checkpoint.mark_failed(chunk_id)
            raise
    """================= End method _flush_sql ================="""

    """================= Startup method _init_excel_workbook ================="""
    def _init_excel_workbook(self, target: ExcelTarget):
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl required for Excel output. pip install openpyxl"
            )

        write_only = target.total_rows > EXCEL_WRITEONLY_THRESHOLD

        if write_only:
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet(title="extracted")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "extracted"

        header = self.schema.final_cols + ["__orig_rowid__"]
        ws.append(header)

        Path(target.output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(target.output_path)

        logger.info(
            "[ReverseWriter] Excel workbook ready: %s  write_only=%s",
            target.output_path, write_only,
        )
        return wb, ws
    """================= End method _init_excel_workbook ================="""

    """================= Startup method _flush_excel ================="""
    def _flush_excel(self, buffer: list[dict], chunk_id: int) -> None:
        target     = self.target
        final_cols = self.schema.final_cols + ["__orig_rowid__"]

        try:
            for row in buffer:
                self._ws.append([row.get(c) for c in final_cols])

            self._wb.save(target.output_path)
            n = len(buffer)
            with self._lock:
                self._total_rows_written += n
            self.checkpoint.mark_done(chunk_id, rows_done=n)
            logger.debug("[ReverseWriter] Excel flush  chunk=%d  rows=%d", chunk_id, n)
        except Exception as exc:
            logger.error("[ReverseWriter] Excel flush error chunk=%d: %s", chunk_id, exc)
            self.checkpoint.mark_failed(chunk_id)
            raise
    """================= End method _flush_excel ================="""

"""================= End class ReverseWriter ================="""


"""================= Startup function make_sql_target ================="""
def make_sql_target(
    table_name: str,
    pg_config:  dict,
) -> SQLTarget:
    """Create a PostgreSQL/Neon SQL target. pg_config is passed directly to psycopg2.connect()."""
    target           = SQLTarget(conn=None, table_name=table_name, dialect="postgresql")
    target._pg_config = pg_config
    return target
"""================= End function make_sql_target ================="""


"""================= Startup function make_excel_target ================="""
def make_excel_target(
    output_path: str,
    total_rows:  int = 0,
) -> ExcelTarget:
    return ExcelTarget(output_path=output_path, total_rows=total_rows)
"""================= End function make_excel_target ================="""